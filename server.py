from __future__ import annotations

import io
import json
import re
import warnings
from dataclasses import dataclass
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse
from typing import Any

warnings.filterwarnings("ignore", category=DeprecationWarning)

import cgi

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
OUTPUT_DIR = ROOT / "outputs"
DATA_DIR = ROOT / "data"
MASTER_PIVOT = DATA_DIR / "master_pivot.xlsx"

BASE_HEADERS = ["약품코드", "처방명", "KD코드", "제약사", "판매사", "특별약품", "단가", "신약여부"]
REQUIRED_SOURCE_FIELDS = BASE_HEADERS[:7] + ["원내불출량"]

ALIASES = {
    "약품코드": ["약품코드", "약품 코드", "품목코드", "품목 코드", "의약품코드", "약제코드", "코드"],
    "처방명": ["처방명", "처방 명", "약품명", "약품 명", "품목명", "제품명", "의약품명", "약명"],
    "KD코드": ["KD코드", "KD 코드", "KD_CODE", "KD CODE", "KD"],
    "제약사": ["제약사", "제약회사", "제약 회사", "제조사", "제조업체", "업체명"],
    "판매사": ["판매사", "판매회사", "판매 회사", "판매업체", "공급사", "유통사"],
    "특별약품": ["특별약품", "특별 약품", "특수약품", "분류", "구분"],
    "단가": ["단가", "단가가", "보험단가", "약가", "금액단가"],
    "원내불출량": ["원내불출량", "원내 불출량", "불출량", "사용량", "수량"],
}


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_()\[\]{}./-]+", "", str(value or "")).upper()


NORMALIZED_ALIASES = {
    field: {normalize_header(alias) for alias in aliases} for field, aliases in ALIASES.items()
}


@dataclass
class SourceParseResult:
    rows: list[dict[str, Any]]
    skipped_without_kd: int
    duplicated_rows: int
    header_row: int
    detected_month: str | None


def parse_number(value: Any) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in {"", "-", ".", "-."}:
        return 0
    try:
        return float(cleaned)
    except ValueError:
        return 0


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_search(value: Any) -> str:
    return re.sub(r"[^0-9A-Z가-힣]+", "", clean_text(value).upper())


def canonical_month(raw: str) -> str:
    text = raw.strip()
    match = re.search(r"(\d{4})\D*(\d{1,2})", text)
    if not match:
        raise ValueError("연월은 예: 2025년 10월 또는 2025-10 형식으로 입력해주세요.")
    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        raise ValueError("월은 1부터 12 사이여야 합니다.")
    return f"{year}년 {month}월"


def month_sort_key(month: str) -> tuple[int, int]:
    match = re.search(r"(\d{4})\D*(\d{1,2})", month)
    if not match:
        return (9999, 99)
    return (int(match.group(1)), int(match.group(2)))


def is_missing_kd(value: Any) -> bool:
    text = clean_text(value)
    return not text or text in {"-", "－", "N/A", "NA", "없음"}


def source_key(row: dict[str, Any]) -> str:
    return f"{clean_text(row.get('약품코드'))}||{clean_text(row.get('KD코드'))}"


def detect_month(ws, header_row: int) -> str | None:
    for row_index in range(1, header_row + 1):
        for col_index in range(1, ws.max_column + 1):
            value = clean_text(ws.cell(row_index, col_index).value)
            if re.search(r"\d{4}\D+\d{1,2}\D*", value):
                try:
                    return canonical_month(value)
                except ValueError:
                    continue
    return None


def find_header_map(ws) -> tuple[int, dict[str, int]]:
    best_row = 0
    best_map: dict[str, int] = {}

    for row_index in range(1, min(ws.max_row, 30) + 1):
        candidate: dict[str, int] = {}
        for col_index in range(1, ws.max_column + 1):
            value = ws.cell(row_index, col_index).value
            normalized = normalize_header(value)
            if not normalized:
                continue
            for field, aliases in NORMALIZED_ALIASES.items():
                if field not in candidate and normalized in aliases:
                    candidate[field] = col_index
        if len(candidate) > len(best_map):
            best_row = row_index
            best_map = candidate

    missing = [field for field in REQUIRED_SOURCE_FIELDS if field not in best_map]
    if missing:
        raise ValueError(f"원본 파일에서 필요한 컬럼을 찾지 못했습니다: {', '.join(missing)}")
    return best_row, best_map


def read_source_workbook(file_bytes: bytes) -> SourceParseResult:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, header_map = find_header_map(ws)
    detected_month = detect_month(ws, header_row)

    merged: dict[str, dict[str, Any]] = {}
    skipped_without_kd = 0
    duplicated_rows = 0
    carried_values: dict[str, Any] = {}

    for row_index in range(header_row + 1, ws.max_row + 1):
        row = {
            field: ws.cell(row_index, col_index).value for field, col_index in header_map.items()
        }
        if not any(row.values()):
            continue
        for carried_field in ["약품코드", "처방명", "KD코드", "제약사", "판매사", "특별약품"]:
            if clean_text(row.get(carried_field)):
                carried_values[carried_field] = row.get(carried_field)
            elif carried_values.get(carried_field):
                row[carried_field] = carried_values[carried_field]

        kd_code = clean_text(row.get("KD코드"))
        if is_missing_kd(kd_code):
            skipped_without_kd += 1
            continue

        normalized = {field: clean_text(row.get(field)) for field in BASE_HEADERS[:6]}
        normalized["단가"] = parse_number(row.get("단가"))
        normalized["원내불출량"] = parse_number(row.get("원내불출량"))

        key = source_key(normalized)
        if key in merged:
            merged[key]["원내불출량"] += normalized["원내불출량"]
            duplicated_rows += 1
            for field in BASE_HEADERS[:7]:
                if not merged[key].get(field) and normalized.get(field):
                    merged[key][field] = normalized[field]
        else:
            merged[key] = normalized

    return SourceParseResult(
        rows=list(merged.values()),
        skipped_without_kd=skipped_without_kd,
        duplicated_rows=duplicated_rows,
        header_row=header_row,
        detected_month=detected_month,
    )


def load_pivot(file_bytes: bytes | None) -> tuple[Workbook, Any, dict[str, int], dict[str, int]]:
    if file_bytes:
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb["피벗데이터"] if "피벗데이터" in wb.sheetnames else wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "피벗데이터"
        ws.append(BASE_HEADERS)

    header_map: dict[str, int] = {}
    for col_index in range(1, ws.max_column + 1):
        value = clean_text(ws.cell(1, col_index).value)
        if value:
            header_map[value] = col_index

    for header in BASE_HEADERS:
        if header not in header_map:
            col_index = ws.max_column + 1
            ws.cell(1, col_index).value = header
            header_map[header] = col_index

    key_to_row: dict[str, int] = {}
    for row_index in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_index, header_map[header]).value for header in BASE_HEADERS[:7]}
        key = source_key(row)
        if key != "||":
            key_to_row[key] = row_index

    return wb, ws, header_map, key_to_row


def apply_month(source_bytes: bytes, pivot_bytes: bytes | None, month_raw: str) -> tuple[bytes, dict[str, Any]]:
    parsed = read_source_workbook(source_bytes)
    month = canonical_month(month_raw) if month_raw.strip() else parsed.detected_month
    if not month:
        raise ValueError("연월을 입력하거나, 원본 엑셀 상단에 예: 2025년 10월 형식의 연월이 있어야 합니다.")
    wb, ws, header_map, key_to_row = load_pivot(pivot_bytes)

    if month not in header_map:
        col_index = ws.max_column + 1
        ws.cell(1, col_index).value = month
        header_map[month] = col_index

    added = 0
    updated = 0

    for row in parsed.rows:
        key = source_key(row)
        row_index = key_to_row.get(key)
        if not row_index:
            row_index = ws.max_row + 1
            key_to_row[key] = row_index
            added += 1
            for header in BASE_HEADERS[:7]:
                ws.cell(row_index, header_map[header]).value = row.get(header)
            ws.cell(row_index, header_map["신약여부"]).value = "신약" if pivot_bytes else ""
        else:
            updated += 1
            for header in BASE_HEADERS[:7]:
                if not ws.cell(row_index, header_map[header]).value and row.get(header) not in ("", None):
                    ws.cell(row_index, header_map[header]).value = row.get(header)
        ws.cell(row_index, header_map[month]).value = row["원내불출량"]

    format_pivot_sheet(ws)
    add_summary_sheet(wb, ws)

    stream = io.BytesIO()
    wb.save(stream)
    stats = {
        "month": month,
        "sourceRows": len(parsed.rows),
        "updatedRows": updated,
        "newRows": added,
        "skippedWithoutKd": parsed.skipped_without_kd,
        "duplicatedRowsMerged": parsed.duplicated_rows,
        "headerRow": parsed.header_row,
    }
    return stream.getvalue(), stats


def apply_many_sources(
    sources: list[tuple[str, bytes]], pivot_bytes: bytes | None, month_raw: str
) -> tuple[bytes, list[dict[str, Any]]]:
    if not sources:
        raise ValueError("월별 원본 엑셀 파일을 선택해주세요.")
    if len(sources) > 1 and month_raw.strip():
        raise ValueError("여러 파일을 한 번에 넣을 때는 연월 입력칸을 비워두세요.")

    current_pivot = pivot_bytes
    stats_log = []
    for filename, source_bytes in sorted(sources, key=lambda item: item[0]):
        current_pivot, stats = apply_month(source_bytes, current_pivot, month_raw)
        stats_log.append({"file": filename, **stats})
    return current_pivot, stats_log


def format_pivot_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "I2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "약품코드": 16,
        "처방명": 34,
        "KD코드": 16,
        "제약사": 18,
        "판매사": 18,
        "특별약품": 14,
        "단가": 12,
        "신약여부": 12,
    }
    for col_index in range(1, ws.max_column + 1):
        header = clean_text(ws.cell(1, col_index).value)
        width = widths.get(header, 14)
        ws.column_dimensions[get_column_letter(col_index)].width = width
        if header == "단가" or re.match(r"\d{4}년 \d{1,2}월", header):
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row_index, col_index).number_format = "#,##0.00"


def add_summary_sheet(wb: Workbook, pivot_ws) -> None:
    if "월별요약" in wb.sheetnames:
        del wb["월별요약"]
    ws = wb.create_sheet("월별요약", 0)

    month_cols = []
    for col_index in range(1, pivot_ws.max_column + 1):
        header = clean_text(pivot_ws.cell(1, col_index).value)
        if re.match(r"\d{4}년 \d{1,2}월", header):
            month_cols.append((header, col_index))

    ws.append(["연월", "총 원내불출량", "품목 수"])
    for month, col_index in month_cols:
        total = 0.0
        item_count = 0
        for row_index in range(2, pivot_ws.max_row + 1):
            value = parse_number(pivot_ws.cell(row_index, col_index).value)
            if value:
                total += value
                item_count += 1
        ws.append([month, total, item_count])

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="70AD47")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    for row_index in range(2, ws.max_row + 1):
        ws.cell(row_index, 2).number_format = "#,##0.00"


def master_exists() -> bool:
    return MASTER_PIVOT.exists()


def read_master_bytes() -> bytes | None:
    return MASTER_PIVOT.read_bytes() if master_exists() else None


def save_master(output_bytes: bytes) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    MASTER_PIVOT.write_bytes(output_bytes)


def pivot_sheet_from_master():
    if not master_exists():
        raise ValueError("아직 기준 피벗 데이터가 없습니다. 먼저 월별 엑셀을 업로드해주세요.")
    wb = load_workbook(MASTER_PIVOT, data_only=True)
    ws = wb["피벗데이터"] if "피벗데이터" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    return wb, ws, headers


def month_headers(headers: list[str]) -> list[str]:
    return sorted(
        [header for header in headers if re.match(r"\d{4}년 \d{1,2}월", header)],
        key=month_sort_key,
    )


def row_to_item(ws, headers: list[str], row_index: int) -> dict[str, Any]:
    item = {"row": row_index, "key": ""}
    for header in BASE_HEADERS:
        if header in headers:
            item[header] = clean_text(ws.cell(row_index, headers.index(header) + 1).value)
        else:
            item[header] = ""
    item["key"] = source_key(item)
    return item


def drug_group_key(item: dict[str, Any]) -> str:
    return f"{normalize_search(item.get('약품코드'))}||{normalize_search(item.get('처방명'))}"


def score_item(query: str, values: list[str]) -> int:
    normalized_query = normalize_search(query)
    if not normalized_query:
        return 0

    best = 0
    query_parts = [normalize_search(part) for part in re.split(r"\s+", query) if part.strip()]
    for value in values:
        normalized_value = normalize_search(value)
        if not normalized_value:
            continue
        if normalized_value == normalized_query:
            best = max(best, 120)
        elif normalized_value.startswith(normalized_query):
            best = max(best, 95)
        elif normalized_query in normalized_value:
            best = max(best, 78)
        elif all(part and part in normalized_value for part in query_parts):
            best = max(best, 68)
        else:
            cursor = 0
            matched = 0
            for char in normalized_query:
                found = normalized_value.find(char, cursor)
                if found < 0:
                    break
                matched += 1
                cursor = found + 1
            if normalized_query:
                best = max(best, int(45 * matched / len(normalized_query)))
    return best


def search_master(query: str, limit: int = 20) -> list[dict[str, Any]]:
    _, ws, headers = pivot_sheet_from_master()
    searchable_headers = [
        header for header in headers if header and not re.match(r"\d{4}년 \d{1,2}월", header)
    ]
    results = []

    for row_index in range(2, ws.max_row + 1):
        values = [
            clean_text(ws.cell(row_index, headers.index(header) + 1).value)
            for header in searchable_headers
        ]
        score = score_item(query, values)
        if score <= 0:
            continue
        item = row_to_item(ws, headers, row_index)
        item["score"] = score
        item["display"] = " / ".join(
            part for part in [item.get("약품코드"), item.get("처방명"), item.get("KD코드")] if part
        )
        results.append(item)

    return sorted(results, key=lambda item: (-item["score"], item["display"]))[:limit]


def item_trend(item_key: str) -> dict[str, Any]:
    _, ws, headers = pivot_sheet_from_master()
    months = month_headers(headers)
    target_item = None

    for row_index in range(2, ws.max_row + 1):
        item = row_to_item(ws, headers, row_index)
        if item["key"] == item_key:
            target_item = item
            break

    if not target_item:
        raise ValueError("선택한 품목을 기준 피벗에서 찾지 못했습니다.")

    target_group = drug_group_key(target_item)
    grouped_rows = []
    for row_index in range(2, ws.max_row + 1):
        item = row_to_item(ws, headers, row_index)
        if drug_group_key(item) == target_group:
            grouped_rows.append((row_index, item))

    components = []
    for row_index, item in grouped_rows:
        component_series = []
        for month in months:
            component_series.append(
                {
                    "month": month,
                    "value": parse_number(ws.cell(row_index, headers.index(month) + 1).value),
                }
            )
        components.append({"item": item, "series": component_series})

    previous = None
    series = []
    for index, month in enumerate(months):
        value = sum(component["series"][index]["value"] for component in components)
        change = None
        if previous not in (None, 0):
            change = ((value - previous) / previous) * 100
        series.append({"month": month, "value": value, "changePct": change})
        previous = value

    return {"item": target_item, "series": series, "components": components}


def master_status() -> dict[str, Any]:
    if not master_exists():
        return {"exists": False, "rows": 0, "months": []}
    _, ws, headers = pivot_sheet_from_master()
    return {
        "exists": True,
        "rows": max(ws.max_row - 1, 0),
        "months": month_headers(headers),
        "updatedAt": datetime.fromtimestamp(MASTER_PIVOT.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    }


class AppHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        if path == "/":
            self.serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
        elif path == "/app.js":
            self.serve_file(STATIC_DIR / "app.js", "text/javascript; charset=utf-8")
        elif path == "/style.css":
            self.serve_file(STATIC_DIR / "style.css", "text/css; charset=utf-8")
        elif path.startswith("/api/"):
            try:
                if path == "/api/status":
                    self.send_json({"ok": True, "status": master_status()})
                elif path == "/api/search":
                    query = parse_qs(parsed_url.query).get("q", [""])[0]
                    self.send_json({"ok": True, "results": search_master(query) if query.strip() else []})
                elif path == "/api/item":
                    item_key = parse_qs(parsed_url.query).get("key", [""])[0]
                    self.send_json({"ok": True, **item_trend(item_key)})
                else:
                    self.send_error(HTTPStatus.NOT_FOUND)
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=400)
        elif path == "/download-master":
            self.serve_download(MASTER_PIVOT, "약품_기준_피벗.xlsx")
        else:
            self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        if path not in {"/process", "/update"}:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            month = form.getfirst("month", "")
            source_items = form["source"] if "source" in form else []
            pivot_item = form["pivot"] if "pivot" in form else None
            if not isinstance(source_items, list):
                source_items = [source_items]

            sources = [
                (item.filename, item.file.read())
                for item in source_items
                if getattr(item, "filename", "") and getattr(item, "file", None)
            ]
            pivot_bytes = None
            if pivot_item is not None and getattr(pivot_item, "filename", ""):
                pivot_bytes = pivot_item.file.read()
            elif path == "/update":
                pivot_bytes = read_master_bytes()

            output_bytes, stats_log = apply_many_sources(sources, pivot_bytes, month)
            if path == "/update":
                save_master(output_bytes)
            first_month = stats_log[0]["month"]
            last_month = stats_log[-1]["month"]
            total_new = sum(item["newRows"] for item in stats_log)
            total_skipped = sum(item["skippedWithoutKd"] for item in stats_log)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            month_label = first_month if first_month == last_month else f"{first_month}_{last_month}"
            output_name = f"약품_월별_피벗_{month_label.replace(' ', '_')}_{timestamp}.xlsx"
            OUTPUT_DIR.mkdir(exist_ok=True)
            (OUTPUT_DIR / output_name).write_bytes(output_bytes)

            payload = {
                "ok": True,
                "filename": output_name,
                "stats": {
                    "month": month_label,
                    "sourceRows": sum(item["sourceRows"] for item in stats_log),
                    "updatedRows": sum(item["updatedRows"] for item in stats_log),
                    "newRows": total_new,
                    "skippedWithoutKd": total_skipped,
                    "duplicatedRowsMerged": sum(item["duplicatedRowsMerged"] for item in stats_log),
                    "headerRow": ", ".join(str(item["headerRow"]) for item in stats_log),
                    "fileCount": len(stats_log),
                },
                "monthStats": stats_log,
                "status": master_status() if path == "/update" else None,
                "file": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                + __import__("base64").b64encode(output_bytes).decode("ascii"),
            }
            self.send_json(payload)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def serve_file(self, path: Path, content_type: str) -> None:
        if not path.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.end_headers()
        self.wfile.write(path.read_bytes())

    def serve_download(self, path: Path, filename: str) -> None:
        if not path.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        data = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    server = ThreadingHTTPServer(("127.0.0.1", 5174), AppHandler)
    print("http://127.0.0.1:5174")
    server.serve_forever()


if __name__ == "__main__":
    main()
