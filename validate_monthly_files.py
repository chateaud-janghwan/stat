from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from server import apply_month, parse_number


INPUT_DIR = Path("/Users/chateaud/Downloads/20260530160717_20240583")
OUTPUT_DIR = Path("outputs")
OUTPUT_XLSX = OUTPUT_DIR / "약품_월별_피벗_202505_202604.xlsx"
OUTPUT_STATS = OUTPUT_DIR / "약품_월별_피벗_202505_202604_stats.json"

EXPECTED_MONTHS = [
    "2025년 5월",
    "2025년 6월",
    "2025년 7월",
    "2025년 8월",
    "2025년 9월",
    "2025년 10월",
    "2025년 11월",
    "2025년 12월",
    "2026년 1월",
    "2026년 2월",
    "2026년 3월",
    "2026년 4월",
]


def main() -> None:
    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if len(files) != len(EXPECTED_MONTHS):
        raise RuntimeError(f"Expected {len(EXPECTED_MONTHS)} files, found {len(files)}")

    pivot_bytes = None
    stats_log = []

    for file_path in files:
        pivot_bytes, stats = apply_month(file_path.read_bytes(), pivot_bytes, "")
        stats_log.append({"file": file_path.name, **stats})
        print(
            f"{file_path.name}: month={stats['month']} source={stats['sourceRows']} "
            f"updated={stats['updatedRows']} new={stats['newRows']} skipped={stats['skippedWithoutKd']}"
        )

    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_XLSX.write_bytes(pivot_bytes)
    OUTPUT_STATS.write_text(json.dumps(stats_log, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = load_workbook(OUTPUT_XLSX, data_only=True)
    ws = wb["피벗데이터"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    missing_months = [month for month in EXPECTED_MONTHS if month not in headers]
    if missing_months:
        raise RuntimeError(f"Missing month columns: {missing_months}")

    month_indexes = {month: headers.index(month) + 1 for month in EXPECTED_MONTHS}
    nonempty_by_month = {}
    total_by_month = {}
    for month, col_index in month_indexes.items():
        nonempty = 0
        total = 0.0
        for row_index in range(2, ws.max_row + 1):
            value = ws.cell(row_index, col_index).value
            if value not in (None, ""):
                nonempty += 1
                total += parse_number(value)
        nonempty_by_month[month] = nonempty
        total_by_month[month] = total

    invalid_kd = []
    kd_col = headers.index("KD코드") + 1
    for row_index in range(2, ws.max_row + 1):
        kd_value = str(ws.cell(row_index, kd_col).value or "").strip()
        if kd_value in {"", "-", "－", "N/A", "NA", "없음"}:
            invalid_kd.append(row_index)

    print(f"OUTPUT={OUTPUT_XLSX}")
    print(f"pivot_rows={ws.max_row - 1}")
    print(f"pivot_cols={ws.max_column}")
    print(f"invalid_kd_rows={len(invalid_kd)}")
    print("nonempty_by_month=" + json.dumps(nonempty_by_month, ensure_ascii=False))
    print("total_by_month=" + json.dumps(total_by_month, ensure_ascii=False))

    if invalid_kd:
        raise RuntimeError(f"Found rows with missing KD code: first rows {invalid_kd[:10]}")


if __name__ == "__main__":
    main()
