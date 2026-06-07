from __future__ import annotations

from openpyxl import load_workbook


wb = load_workbook("data/master_pivot.xlsx", data_only=True, read_only=True)
ws = wb["피벗데이터"]
headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
months = [header for header in headers if isinstance(header, str) and "년" in header and "월" in header]
rows = []

for row in ws.iter_rows(min_row=2, values_only=True):
    item = dict(zip(headers, row))
    if "pantoline" in str(item.get("처방명", "")).lower():
        rows.append(item)

print("rows", len(rows))
for item in rows:
    print(item["약품코드"], item["처방명"], item.get("2026년 4월"), [item.get(month) for month in months])
print("sum 202604", sum(float(item.get("2026년 4월") or 0) for item in rows))
