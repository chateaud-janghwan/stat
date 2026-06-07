from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


path = Path("/Users/chateaud/Downloads/20260530160717_20240583/202604약품사용현황(단가포함).xlsx")
needle = ""
wb = load_workbook(path, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]

for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
    values = list(row)
    if row_index == 2:
        print(values)
    if 552 <= row_index <= 565:
        print(row_index, values)
